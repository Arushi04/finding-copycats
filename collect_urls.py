import argparse
import random
import requests
import os

import pandas as pd
from openpyxl import load_workbook

'''
Steps to Run:
python collect_urls.py --start 1 --stop 10339999 --max_urls 10000 --filename urls.csv
'''


def append_df_to_excel(filename, df, sheet_name='Valid_URLs', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        writer.book = load_workbook(filename)       # try to open an existing workbook

        # get the last row in the existing Excel sheet if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:     # truncate sheet
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        print("File not found")

    df.to_excel(writer, sheet_name, header=False, startrow=startrow, **to_excel_kwargs)
    writer.save()


def save_to_excel(num, url, filename, sheetname):
    #create a dataframe
    df = pd.DataFrame({'Numbers': [num], 'URLs': [url]})
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    if os.path.isfile(filename):
        required_df = pd.read_excel(filename, usecols=[0])  # load the file to a dataframe
        if num not in required_df.Numbers.tolist():     # If number is not in the file, append it
            append_df_to_excel(filename, df, sheet_name=sheetname, index=False)
    else:
        df.to_excel(writer, sheetname, index=False)
        writer.save()


def generate_url(start, stop):
    #  generate a random number between start and stop
    num = random.randrange(start, stop)

    #  Append the number to the url
    url = 'https://us.vestiairecollective.com/members/profile-' + str(num) + '.shtml'
    return num, url


def main(args):
    #  creating session to check validity of url
    session = requests.Session()
    num_set = set()  # set to avoid duplication of url
    random.seed(args.seed)      # setting seed so that each time program gives same result

    while len(num_set) < args.max_urls:
        num, url = generate_url(args.start, args.stop)
        if num not in num_set:
            status_code = session.get(url, headers={'User-Agent': 'Mozilla/5.0'}).status_code
            if status_code == 200:  # check if url is valid
                print(url)
                num_set.add(num)
                save_to_excel(num, url, args.filename, args.sheetname)

    print(f"Writing {len(num_set)} urls to csv {args.filename}")



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Arguments')
    parser.add_argument("--start", type=int, default=1, help='number starts generating from this number')
    parser.add_argument("--stop", type=int, default=10339999, help='number generates upto this number')
    parser.add_argument("--max_urls", type=int, default=10000, help='total no of urls we want to create')
    parser.add_argument("--filename", type=str, default='valid_urls.xlsx', help='name of the excel file')
    parser.add_argument("--sheetname", type=str, default='Valid_URLs', help='name of the sheet in excel')
    parser.add_argument("--seed", type=str, default=4, help='to ensure random number is always generated in same sequence')
    args = parser.parse_args()
    main(args)
